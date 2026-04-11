from __future__ import annotations

import json
from datetime import datetime
from io import BytesIO
from typing import Any, Mapping
from urllib.parse import quote

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

MAX_TESTCASE_EXPORT_ROWS = 10000
_EXPORT_MEDIA_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

DEFAULT_TESTCASE_EXPORT_COLUMNS = [
    "case_id",
    "title",
    "module_name",
    "priority",
    "status",
    "description",
    "preconditions",
    "steps",
    "expected_results",
    "test_type",
    "design_technique",
    "test_data",
    "remarks",
    "quality_gate",
    "quality_score",
    "batch_id",
    "source_document_ids",
    "created_at",
    "updated_at",
]


def _coerce_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _coerce_mapping(value: Any) -> dict[str, Any]:
    return dict(value) if isinstance(value, Mapping) else {}


def _coerce_list(value: Any) -> list[Any]:
    return list(value) if isinstance(value, list) else []


def _join_lines(value: Any) -> str:
    items = []
    for item in _coerce_list(value):
        text = _coerce_text(item)
        if text:
            items.append(text)
    return "\n".join(items)


def _json_dump(value: Any) -> str:
    if value in (None, "", [], {}):
        return ""
    return json.dumps(value, ensure_ascii=False, indent=2, default=str)


def _content_json(case: Mapping[str, Any]) -> dict[str, Any]:
    return _coerce_mapping(case.get("content_json"))


def _content_meta(case: Mapping[str, Any]) -> dict[str, Any]:
    return _coerce_mapping(_content_json(case).get("meta"))


def _field(case: Mapping[str, Any], key: str) -> Any:
    if key in case and case.get(key) not in (None, ""):
        return case.get(key)
    return _content_json(case).get(key)


def _quality_review(case: Mapping[str, Any]) -> dict[str, Any]:
    meta = _content_meta(case)
    return _coerce_mapping(meta.get("quality_review"))


def _column_value(case: Mapping[str, Any], column: str) -> str:
    quality_review = _quality_review(case)
    mapping: dict[str, Any] = {
        "case_id": _coerce_text(_field(case, "case_id")),
        "title": _coerce_text(_field(case, "title")),
        "module_name": _coerce_text(_field(case, "module_name")),
        "priority": _coerce_text(_field(case, "priority")),
        "status": _coerce_text(_field(case, "status") or case.get("status")),
        "description": _coerce_text(_field(case, "description") or case.get("description")),
        "preconditions": _join_lines(_field(case, "preconditions")),
        "steps": _join_lines(_field(case, "steps")),
        "expected_results": _join_lines(_field(case, "expected_results")),
        "test_type": _coerce_text(_field(case, "test_type")),
        "design_technique": _coerce_text(_field(case, "design_technique")),
        "test_data": _json_dump(_field(case, "test_data")),
        "remarks": _coerce_text(_field(case, "remarks")),
        "quality_gate": _coerce_text(quality_review.get("quality_gate")),
        "quality_score": _coerce_text(quality_review.get("quality_score")),
        "batch_id": _coerce_text(case.get("batch_id")),
        "source_document_ids": _join_lines(
            case.get("source_document_ids") or _content_meta(case).get("source_document_ids")
        ),
        "created_at": _coerce_text(case.get("created_at")),
        "updated_at": _coerce_text(case.get("updated_at")),
    }
    return _coerce_text(mapping.get(column))


TESTCASE_EXPORT_COLUMN_DEFS: dict[str, tuple[str, int]] = {
    "case_id": ("Case ID", 18),
    "title": ("标题", 28),
    "module_name": ("模块", 18),
    "priority": ("优先级", 10),
    "status": ("状态", 12),
    "description": ("描述", 28),
    "preconditions": ("前置条件", 20),
    "steps": ("步骤", 42),
    "expected_results": ("预期结果", 42),
    "test_type": ("测试类型", 14),
    "design_technique": ("设计技术", 16),
    "test_data": ("测试数据", 24),
    "remarks": ("备注", 18),
    "quality_gate": ("质量门禁", 12),
    "quality_score": ("质量分数", 12),
    "batch_id": ("批次 ID", 38),
    "source_document_ids": ("来源文档 IDs", 32),
    "created_at": ("创建时间", 22),
    "updated_at": ("更新时间", 22),
}


def resolve_testcase_export_columns(columns: list[str] | None) -> list[str]:
    if not columns:
        return list(DEFAULT_TESTCASE_EXPORT_COLUMNS)
    resolved: list[str] = []
    for item in columns:
        key = _coerce_text(item)
        if key and key in TESTCASE_EXPORT_COLUMN_DEFS and key not in resolved:
            resolved.append(key)
    return resolved or list(DEFAULT_TESTCASE_EXPORT_COLUMNS)


def _export_filename(project_id: str) -> str:
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    project_key = (project_id or "project")[:8]
    return f"testcase-cases-{project_key}-{timestamp}.xlsx"


def build_content_disposition(filename: str) -> str:
    encoded = quote(filename)
    return f"attachment; filename=\"{filename}\"; filename*=UTF-8''{encoded}"


def _styled_row(
    sheet,
    values: list[str],
    *,
    font: Font | None = None,
    fill: PatternFill | None = None,
    alignment: Alignment | None = None,
) -> list[WriteOnlyCell]:
    row: list[WriteOnlyCell] = []
    for value in values:
        cell = WriteOnlyCell(sheet, value=value)
        if font is not None:
            cell.font = font
        if fill is not None:
            cell.fill = fill
        if alignment is not None:
            cell.alignment = alignment
        row.append(cell)
    return row


def build_testcase_cases_workbook(
    *,
    project_id: str,
    cases: list[dict[str, Any]],
    filters: Mapping[str, Any],
    columns: list[str] | None = None,
) -> tuple[str, bytes]:
    resolved_columns = resolve_testcase_export_columns(columns)
    workbook = Workbook(write_only=True)
    cases_sheet = workbook.create_sheet("测试用例")
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    wrap_alignment = Alignment(vertical="top", wrap_text=True)
    widths = [TESTCASE_EXPORT_COLUMN_DEFS[column][1] for column in resolved_columns]
    for index, width in enumerate(widths, start=1):
        cases_sheet.column_dimensions[get_column_letter(index)].width = width
    cases_sheet.freeze_panes = "A2"
    cases_sheet.auto_filter.ref = f"A1:{get_column_letter(len(resolved_columns))}{max(len(cases) + 1, 1)}"

    headers = [TESTCASE_EXPORT_COLUMN_DEFS[column][0] for column in resolved_columns]
    cases_sheet.append(
        _styled_row(
            cases_sheet,
            headers,
            font=header_font,
            fill=header_fill,
            alignment=wrap_alignment,
        )
    )

    for case in cases:
        row = [_column_value(case, column) for column in resolved_columns]
        cases_sheet.append(
            _styled_row(
                cases_sheet,
                row,
                alignment=wrap_alignment,
            )
        )

    meta_sheet = workbook.create_sheet("导出信息")
    meta_sheet.append(
        _styled_row(
            meta_sheet,
            ["字段", "值"],
            font=header_font,
            fill=header_fill,
            alignment=wrap_alignment,
        )
    )
    meta_sheet.append(_styled_row(meta_sheet, ["project_id", project_id], alignment=wrap_alignment))
    meta_sheet.append(
        _styled_row(
            meta_sheet,
            ["导出时间", datetime.now().isoformat(timespec="seconds")],
            alignment=wrap_alignment,
        )
    )
    meta_sheet.append(_styled_row(meta_sheet, ["导出总条数", str(len(cases))], alignment=wrap_alignment))
    meta_sheet.append(
        _styled_row(
            meta_sheet,
            ["batch_id", _coerce_text(filters.get("batch_id"))],
            alignment=wrap_alignment,
        )
    )
    meta_sheet.append(
        _styled_row(
            meta_sheet,
            ["status", _coerce_text(filters.get("status"))],
            alignment=wrap_alignment,
        )
    )
    meta_sheet.append(
        _styled_row(
            meta_sheet,
            ["query", _coerce_text(filters.get("query"))],
            alignment=wrap_alignment,
        )
    )
    meta_sheet.append(
        _styled_row(
            meta_sheet,
            ["columns", ",".join(resolved_columns)],
            alignment=wrap_alignment,
        )
    )
    meta_sheet.column_dimensions["A"].width = 18
    meta_sheet.column_dimensions["B"].width = 48

    output = BytesIO()
    workbook.save(output)
    return _export_filename(project_id), output.getvalue()


__all__ = [
    "MAX_TESTCASE_EXPORT_ROWS",
    "_EXPORT_MEDIA_TYPE",
    "DEFAULT_TESTCASE_EXPORT_COLUMNS",
    "TESTCASE_EXPORT_COLUMN_DEFS",
    "build_content_disposition",
    "build_testcase_cases_workbook",
    "resolve_testcase_export_columns",
]
