from __future__ import annotations

import sys
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

_PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(_PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(_PROJECT_ROOT))

from app.services.testcase_case_export import (
    DEFAULT_TESTCASE_EXPORT_COLUMNS,
    build_testcase_cases_workbook,
    resolve_testcase_export_columns,
)
from app.services.testcase_document_export import build_testcase_documents_workbook


def test_resolve_testcase_export_columns_keeps_whitelist_and_order() -> None:
    resolved = resolve_testcase_export_columns(
        ["title", "steps", "title", "unknown", "updated_at", ""]
    )
    assert resolved == ["title", "steps", "updated_at"]
    assert resolve_testcase_export_columns([]) == DEFAULT_TESTCASE_EXPORT_COLUMNS


def test_build_testcase_cases_workbook_supports_custom_columns() -> None:
    filename, payload = build_testcase_cases_workbook(
        project_id="5f419550-a3c7-49c6-9450-09154fd1bf7d",
        cases=[
            {
                "id": "case-1",
                "case_id": "TC-LOGIN-001",
                "title": "登录成功",
                "status": "active",
                "batch_id": "test-case-service:batch-1",
                "source_document_ids": ["doc-1", "doc-2"],
                "content_json": {
                    "steps": ["输入账号密码", "点击登录"],
                    "expected_results": ["登录成功"],
                    "meta": {"quality_review": {"quality_gate": "pass", "quality_score": 96}},
                },
                "updated_at": "2026-03-31T10:00:00",
            }
        ],
        filters={"batch_id": "test-case-service:batch-1", "status": "active", "query": "登录"},
        columns=["title", "steps", "quality_gate", "source_document_ids"],
    )

    assert filename.startswith("testcase-cases-")

    workbook = load_workbook(BytesIO(payload))
    cases_sheet = workbook["测试用例"]
    meta_sheet = workbook["导出信息"]

    headers = [cell.value for cell in cases_sheet[1]]
    assert headers == ["标题", "步骤", "质量门禁", "来源文档 IDs"]
    row = [cases_sheet.cell(row=2, column=index).value for index in range(1, 5)]
    assert row == ["登录成功", "输入账号密码\n点击登录", "pass", "doc-1\ndoc-2"]
    assert meta_sheet["B5"].value == "test-case-service:batch-1"
    assert meta_sheet["B6"].value == "active"
    assert meta_sheet["B7"].value == "登录"
    assert meta_sheet["B8"].value == "title,steps,quality_gate,source_document_ids"


def test_build_testcase_documents_workbook_includes_runtime_and_asset_fields() -> None:
    filename, payload = build_testcase_documents_workbook(
        project_id="5f419550-a3c7-49c6-9450-09154fd1bf7d",
        documents=[
            {
                "id": "doc-1",
                "filename": "接口文档.pdf",
                "content_type": "application/pdf",
                "source_kind": "upload",
                "parse_status": "parsed",
                "batch_id": "test-case-service:batch-1",
                "summary_for_model": "提取出登录接口关键规则",
                "parsed_text": "登录接口要求用户名密码必填。",
                "structured_data": {"key_points": ["login"]},
                "provenance": {
                    "runtime": {
                        "thread_id": "thread-123",
                        "run_id": "run-456",
                        "agent_key": "test_case_service",
                    },
                    "asset": {"storage_path": "test-case-service/project/default/doc-1.pdf"},
                },
                "related_cases_count": 2,
                "created_at": "2026-03-31T10:00:00",
            }
        ],
        filters={"batch_id": "test-case-service:batch-1", "parse_status": "parsed", "query": "登录"},
    )

    assert filename.startswith("testcase-documents-")

    workbook = load_workbook(BytesIO(payload))
    sheet = workbook["PDF解析记录"]
    meta_sheet = workbook["导出信息"]

    headers = [cell.value for cell in sheet[1]]
    assert headers == [
        "Document ID",
        "文件名",
        "content_type",
        "source_kind",
        "parse_status",
        "batch_id",
        "summary_for_model",
        "parsed_text_excerpt",
        "parsed_text_length",
        "structured_data_json",
        "thread_id",
        "run_id",
        "agent_key",
        "storage_path",
        "related_cases_count",
        "created_at",
    ]
    row = [sheet.cell(row=2, column=index).value for index in range(1, 17)]
    assert row[0] == "doc-1"
    assert row[10:14] == [
        "thread-123",
        "run-456",
        "test_case_service",
        "test-case-service/project/default/doc-1.pdf",
    ]
    assert row[14] == "2"
    assert meta_sheet["B5"].value == "test-case-service:batch-1"
    assert meta_sheet["B6"].value == "parsed"
    assert meta_sheet["B7"].value == "登录"
