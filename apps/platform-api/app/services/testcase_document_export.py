from __future__ import annotations

import json
from datetime import datetime
from io import BytesIO
from typing import Any, Mapping
from urllib.parse import quote

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

MAX_TESTCASE_DOCUMENT_EXPORT_ROWS = 10000
TESTCASE_DOCUMENT_EXPORT_MEDIA_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


def _coerce_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _coerce_mapping(value: Any) -> dict[str, Any]:
    return dict(value) if isinstance(value, Mapping) else {}


def _json_dump(value: Any) -> str:
    if value in (None, "", [], {}):
        return ""
    return json.dumps(value, ensure_ascii=False, indent=2, default=str)


def _runtime_meta(document: Mapping[str, Any]) -> dict[str, Any]:
    provenance = _coerce_mapping(document.get("provenance"))
    return _coerce_mapping(provenance.get("runtime"))


def _asset_meta(document: Mapping[str, Any]) -> dict[str, Any]:
    provenance = _coerce_mapping(document.get("provenance"))
    return _coerce_mapping(provenance.get("asset"))


def _parsed_text_excerpt(document: Mapping[str, Any], *, max_chars: int = 8000) -> str:
    parsed_text = _coerce_text(document.get("parsed_text"))
    if not parsed_text:
        return ""
    return parsed_text[:max_chars]


def _export_filename(project_id: str) -> str:
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    project_key = (project_id or "project")[:8]
    return f"testcase-documents-{project_key}-{timestamp}.xlsx"


def build_content_disposition(filename: str) -> str:
    encoded = quote(filename)
    return f"attachment; filename=\"{filename}\"; filename*=UTF-8''{encoded}"


def _styled_row(
    sheet,
    values: list[str | int],
    *,
    font: Font | None = None,
    fill: PatternFill | None = None,
    alignment: Alignment | None = None,
) -> list[WriteOnlyCell]:
    row: list[WriteOnlyCell] = []
    for value in values:
        cell = WriteOnlyCell(sheet, value=value)
        if font is not None:
            cell.font = font
        if fill is not None:
            cell.fill = fill
        if alignment is not None:
            cell.alignment = alignment
        row.append(cell)
    return row


def build_testcase_documents_workbook(
    *,
    project_id: str,
    documents: list[dict[str, Any]],
    filters: Mapping[str, Any],
) -> tuple[str, bytes]:
    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet("PDF解析记录")
    headers = [
        "Document ID",
        "文件名",
        "content_type",
        "source_kind",
        "parse_status",
        "batch_id",
        "summary_for_model",
        "parsed_text_excerpt",
        "parsed_text_length",
        "structured_data_json",
        "thread_id",
        "run_id",
        "agent_key",
        "storage_path",
        "related_cases_count",
        "created_at",
    ]
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    wrap_alignment = Alignment(vertical="top", wrap_text=True)
    widths = {
        1: 38,
        2: 28,
        3: 22,
        4: 16,
        5: 14,
        6: 36,
        7: 42,
        8: 60,
        9: 16,
        10: 42,
        11: 28,
        12: 28,
        13: 18,
        14: 44,
        15: 16,
        16: 24,
    }
    for index, width in widths.items():
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(len(documents) + 1, 1)}"
    sheet.append(
        _styled_row(
            sheet,
            headers,
            font=header_font,
            fill=header_fill,
            alignment=wrap_alignment,
        )
    )

    for document in documents:
        runtime_meta = _runtime_meta(document)
        parsed_text = _coerce_text(document.get("parsed_text"))
        row = [
            _coerce_text(document.get("id")),
            _coerce_text(document.get("filename")),
            _coerce_text(document.get("content_type")),
            _coerce_text(document.get("source_kind")),
            _coerce_text(document.get("parse_status")),
            _coerce_text(document.get("batch_id")),
            _coerce_text(document.get("summary_for_model")),
            _parsed_text_excerpt(document),
            len(parsed_text),
            _json_dump(document.get("structured_data")),
            _coerce_text(runtime_meta.get("thread_id")),
            _coerce_text(runtime_meta.get("run_id")),
            _coerce_text(runtime_meta.get("agent_key")),
            _coerce_text(document.get("storage_path") or _asset_meta(document).get("storage_path")),
            _coerce_text(document.get("related_cases_count")),
            _coerce_text(document.get("created_at")),
        ]
        sheet.append(_styled_row(sheet, row, alignment=wrap_alignment))

    meta_sheet = workbook.create_sheet("导出信息")
    meta_sheet.append(
        _styled_row(
            meta_sheet,
            ["字段", "值"],
            font=header_font,
            fill=header_fill,
            alignment=wrap_alignment,
        )
    )
    meta_sheet.append(_styled_row(meta_sheet, ["project_id", project_id], alignment=wrap_alignment))
    meta_sheet.append(
        _styled_row(
            meta_sheet,
            ["导出时间", datetime.now().isoformat(timespec="seconds")],
            alignment=wrap_alignment,
        )
    )
    meta_sheet.append(
        _styled_row(meta_sheet, ["导出总条数", str(len(documents))], alignment=wrap_alignment)
    )
    meta_sheet.append(
        _styled_row(
            meta_sheet,
            ["batch_id", _coerce_text(filters.get("batch_id"))],
            alignment=wrap_alignment,
        )
    )
    meta_sheet.append(
        _styled_row(
            meta_sheet,
            ["parse_status", _coerce_text(filters.get("parse_status"))],
            alignment=wrap_alignment,
        )
    )
    meta_sheet.append(
        _styled_row(
            meta_sheet,
            ["query", _coerce_text(filters.get("query"))],
            alignment=wrap_alignment,
        )
    )
    meta_sheet.column_dimensions["A"].width = 18
    meta_sheet.column_dimensions["B"].width = 48

    output = BytesIO()
    workbook.save(output)
    return _export_filename(project_id), output.getvalue()


__all__ = [
    "MAX_TESTCASE_DOCUMENT_EXPORT_ROWS",
    "TESTCASE_DOCUMENT_EXPORT_MEDIA_TYPE",
    "build_content_disposition",
    "build_testcase_documents_workbook",
]
